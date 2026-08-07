#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extractor de datos de libros para publicaciones de Mercado Libre (versión escritorio).

A partir de FOTOS de libros (tapa, y opcionalmente contratapa / hoja con el ISBN),
usa un modelo de IA con visión + bases de datos gratuitas (Google Books / OpenLibrary)
y genera un Excel con las 61 columnas de la planilla oficial
"Publicar varios productos" de Mercado Libre (Libros Físicos).

Uso rápido:
    1) Copiá las fotos dentro de la carpeta "fotos/"  (una foto por libro).
    2) Configurá tu API key en config.json (ver config.example.json).
    3) Ejecutá:   python extractor.py
    4) Abrí "salida.xlsx" y pegá las filas en la planilla que bajás de Mercado Libre.

Varias fotos del MISMO libro:
    Nombrá los archivos con el mismo prefijo y doble guion bajo, p. ej.:
        ishiguro__tapa.jpg   ishiguro__contra.jpg   ishiguro__isbn.jpg
    Todas las fotos con el prefijo "ishiguro" se combinan en un solo libro.

También:
    python extractor.py --texto "Restos del dia Kazuo Ishiguro"
    python extractor.py --isbn 9788483462287

Ver README.md para la guía paso a paso.
"""

import argparse
import base64
import io
import json
import mimetypes
import os
import re
import sys
import time
from collections import OrderedDict, defaultdict
from pathlib import Path

try:
    import requests
except ImportError:
    print("Falta la librería 'requests'. Instalá con:  pip install -r requirements.txt")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Falta la librería 'openpyxl'. Instalá con:  pip install -r requirements.txt")
    sys.exit(1)

try:
    from PIL import Image
    HAY_PILLOW = True
except ImportError:
    HAY_PILLOW = False


RAIZ = Path(__file__).resolve().parent

CONFIG_POR_DEFECTO = {
    "proveedor": "gemini",              # gemini | openai | anthropic
    "api_key": "",
    "modelo": "",                       # vacío = autodetección (Gemini)
    "carpeta_fotos": "fotos",
    "archivo_salida": "salida.xlsx",
    "etiqueta_titulo": "Microcentro",
    "condicion": "Usado",
    "sku_por_defecto": "",
    "stock": 1,
    "precio": "manual",                 # manual | sugerir | <número>
    "max_largo_titulo": 60,
    # Medidas y peso del PAQUETE de envío (obligatorio en ML).
    "paq_ancho": "24", "paq_alto": "17", "paq_prof": "4", "paq_peso": "0.5",
    # Condiciones de venta.
    "forma_envio": "Mercado Envíos",
    "costo_envio": "A cargo del comprador",
    "retiro": "Acepto",
    "cuotas": "No agregar cuotas",
    "costo_cuotas": "Sin costo",
    # Cierre de la descripción (tu librería).
    "libreria": "Librería Los Siete Pilares",
    "retiro_texto": "Retiro en persona en una librería a la calle en la zona de Paraguay y Reconquista.",
    "titulo_italica": False,
    # Si apuntás a la planilla oficial de ML, se rellena esa (en vez de crear salida.xlsx).
    "plantilla": "",
}

MODELOS_POR_DEFECTO = {
    "gemini": "gemini-2.5-flash",  # si no existe, se autodetecta uno disponible
    "openai": "gpt-4o-mini",
    "anthropic": "claude-3-5-sonnet-20241022",
}
VAR_ENTORNO_KEY = {
    "gemini": "GEMINI_API_KEY",
    "openai": "OPENAI_API_KEY",
    "anthropic": "ANTHROPIC_API_KEY",
}
EXTENSIONES_IMAGEN = {".jpg", ".jpeg", ".png", ".webp", ".heic", ".bmp", ".gif"}

# Las 61 columnas de la planilla, en orden. Cada entrada: (encabezado, clave o None).
# clave None = columna gris/automática (queda vacía).
COLUMNAS = [
    ("Código de catálogo ML", None),
    ("Título", "__titulo"),
    ("Cantidad de caracteres", None),
    ("Condición", "__condicion"),
    ("ISBN", "isbn"),
    ("Fotos", None),
    ("SKU", "__sku"),
    ("Stock", "__stock"),
    ("Precio [$]", "__precio"),
    ("Descripción", "__descripcion"),
    ("Ancho (cm)", "__paq_ancho"),
    ("Alto (cm)", "__paq_alto"),
    ("Profundidad (cm)", "__paq_prof"),
    ("Peso físico (kg)", "__paq_peso"),
    ("Cargo por vender", None),
    ("Cuotas", "__cuotas"),
    ("Costo por ofrecer cuotas", "__costo_cuotas"),
    ("Forma de envío", "__forma_envio"),
    ("Costo de envío", "__costo_envio"),
    ("Retiro en persona", "__retiro"),
    ("Tipo de garantía", None),
    ("Tiempo de garantía", None),
    ("Unidad de Tiempo de garantía", None),
    ("Factura A", None),
    ("Título del libro", "titulo_libro"),
    ("Autor", "autor"),
    ("Editorial del libro", "editorial"),
    ("Subtítulo del libro", "subtitulo"),
    ("Serie", "serie"),
    ("Idioma", "idioma"),
    ("Edición del libro", "edicion"),
    ("Tapa del libro", "tapa"),
    ("Volumen del libro", None),
    ("Índice", "indice"),
    ("Año de publicación", "anio"),
    ("Páginas para colorear", None),
    ("Realidad aumentada", None),
    ("Coautores", "coautores"),
    ("Traductores", "traductores"),
    ("Tipo de narración", "tipo_narracion"),
    ("Versión del libro", None),
    ("BISAC", None),
    ("Tamaño del libro", None),
    ("NCM", None),
    ("País de origen", "pais_origen"),
    ("Colección del libro", "coleccion"),
    ("Edad mínima recomendada", None),
    ("Unidad de Edad mínima recomendada", None),
    ("En imprenta mayúscula", None),
    ("Cantidad de libros por set", None),
    ("Cantidad de páginas", "paginas"),
    ("Altura", "altura_cm"),
    ("Unidad de Altura", "__u_altura"),
    ("Ancho", "ancho_cm"),
    ("Unidad de Ancho", "__u_ancho"),
    ("Peso", "peso_g"),
    ("Unidad de Peso", "__u_peso"),
    ("Material de la tapa del libro", "material_tapa"),
    ("Resumen de errores", None),
    ("BUYBOX_FORMULA", None),
    ("HIDDEN_PICTURES", None),
]

# ---------------------------------------------------------------------------
# Prompt
# ---------------------------------------------------------------------------

ESQUEMA = ('{"titulo_libro":"","subtitulo":"","autor":"","coautores":"","traductores":"",'
           '"editorial":"","coleccion":"","serie":"","edicion":"","idioma":"","tapa":"",'
           '"indice":"","anio":"","isbn":"","tipo_narracion":"","paginas":"","tema_genero":"","ciudad":"",'
           '"pais_origen":"","altura_cm":"","ancho_cm":"","grosor_cm":"","peso_g":"","material_tapa":"",'
           '"condicion":"","estado":"","descripcion":"","precio_sugerido_min":null,'
           '"precio_sugerido_max":null,"confianza":"alta|media|baja","observaciones":""}')

REGLAS = """Reglas de formato:
- "idioma": nombre en español (Español, Inglés, Italiano, Ruso, Francés, Alemán, Portugués, Latín, Griego, etc.). El idioma puede NO ser español: leé el alfabeto real.
- "tapa": exactamente "Dura" o "Blanda" (o "" si no se sabe).
- "indice": "Sí" o "No" (o "").
- "tipo_narracion": p. ej. Novela, Cuento, Poesía, Ensayo, Teatro, Antología (o "").
- "condicion": "Usado" salvo que claramente sea nuevo.
- "estado": frase corta del estado del ejemplar: "Buen estado", "Muy buen estado", "Excelente estado", "Usado con marcas de uso", etc.
- "ciudad": ciudad de edición (donde se editó/publicó), si aparece o la conocés (ej. Barcelona, Buenos Aires).
- "isbn": si aparece en alguna foto (contratapa / hoja de créditos), transcribilo (solo números).
- "paginas": solo número. "peso_g": peso del libro en gramos; NO se puede medir por foto (dejalo "" salvo estimación razonable).
- MEDICIÓN CON REGLA: si en la foto hay una regla u objeto de tamaño conocido (tarjeta = 8.5 cm, moneda, etc.), usalo como escala para medir "altura_cm" (alto), "ancho_cm" (ancho) y "grosor_cm" (grosor del lomo) del libro, en cm. Si no hay referencia, estimá solo si es razonable; si no, "". Indicá en "observaciones" qué referencia usaste.
- "descripcion": 2 a 4 oraciones en español neutro (de qué trata + datos del ejemplar), sin exagerar.
- "precio_sugerido_min"/"precio_sugerido_max": rango orientativo en pesos argentinos para un usado en ML (enteros). SOLO referencia.
- Si algo está borroso o no se ve: dejá "" y bajá la "confianza". NO inventes."""

INSTRUCCIONES = (
    "Sos un experto en catalogación de libros (usados, antiguos y raros) para vender en "
    "Mercado Libre Argentina. Puede que recibas VARIAS fotos del MISMO libro (tapa, contratapa, "
    "hoja de créditos con el ISBN): combiná toda la información en un solo objeto.\n"
    "Devolvé EXCLUSIVAMENTE un JSON válido con EXACTAMENTE estas claves (usá \"\" cuando no tengas el dato):\n"
    + ESQUEMA + "\n" + REGLAS + "\nNo agregues nada fuera del JSON."
)


def prompt_texto(dato=""):
    p = INSTRUCCIONES
    if dato:
        p += f"\nNo hay foto. Identificá el libro a partir de: \"{dato}\""
    return p


# ---------------------------------------------------------------------------
# Imágenes
# ---------------------------------------------------------------------------

def imagen_a_b64(ruta: Path):
    mime, _ = mimetypes.guess_type(str(ruta))
    mime = mime or "image/jpeg"
    if HAY_PILLOW:
        try:
            img = Image.open(ruta).convert("RGB")
            maximo = 1600
            if max(img.size) > maximo:
                e = maximo / max(img.size)
                img = img.resize((int(img.size[0] * e), int(img.size[1] * e)), Image.LANCZOS)
            buf = io.BytesIO(); img.save(buf, format="JPEG", quality=85)
            return base64.b64encode(buf.getvalue()).decode("ascii"), "image/jpeg"
        except Exception:
            pass
    with open(ruta, "rb") as f:
        return base64.b64encode(f.read()).decode("ascii"), mime


# ---------------------------------------------------------------------------
# Proveedores de IA (aceptan una lista de imágenes)
# ---------------------------------------------------------------------------

_gemini_modelos = None
_gemini_malos = set()


def _puntaje_modelo(n):
    s = 0
    if re.search(r"flash", n, re.I): s += 100
    if re.search(r"latest", n, re.I): s += 60
    if re.search(r"2\.0", n): s += 15
    if re.search(r"lite", n, re.I): s -= 25
    if re.search(r"exp|preview|thinking|tts|image|embedding|vision|aqa|learnlm|gemma", n, re.I): s -= 300
    if re.search(r"1\.5|1\.0|8b", n): s -= 120
    return s


def _listar_modelos_gemini(cfg):
    r = requests.get("https://generativelanguage.googleapis.com/v1beta/models",
                     params={"key": cfg["api_key"]}, timeout=60)
    r.raise_for_status()
    nombres = [m["name"].replace("models/", "")
               for m in r.json().get("models", [])
               if "generateContent" in m.get("supportedGenerationMethods", [])]
    return sorted(nombres, key=_puntaje_modelo, reverse=True)


def _pedir_gemini(cfg, modelo, prompt, imagenes):
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{modelo}:generateContent"
    parts = [{"text": prompt}]
    for b64, mime in imagenes:
        parts.append({"inline_data": {"mime_type": mime, "data": b64}})
    return requests.post(url, params={"key": cfg["api_key"]},
                         json={"contents": [{"parts": parts}],
                               "generationConfig": {"response_mime_type": "application/json", "temperature": 0.1}},
                         timeout=180)


def llamar_gemini(cfg, prompt, imagenes):
    global _gemini_modelos
    if cfg.get("modelo"):
        r = _pedir_gemini(cfg, cfg["modelo"], prompt, imagenes)
        if r.ok:
            return r.json()["candidates"][0]["content"]["parts"][0]["text"]
        if r.status_code != 404:
            r.raise_for_status()
    if _gemini_modelos is None:
        _gemini_modelos = _listar_modelos_gemini(cfg)
    for modelo in [m for m in _gemini_modelos if m not in _gemini_malos]:
        r = _pedir_gemini(cfg, modelo, prompt, imagenes)
        if r.ok:
            return r.json()["candidates"][0]["content"]["parts"][0]["text"]
        if r.status_code == 404:
            _gemini_malos.add(modelo); continue
        r.raise_for_status()
    raise RuntimeError("Ningún modelo de Gemini de tu clave está disponible (404). "
                       "Generá la clave en un proyecto nuevo, o usá OpenAI/Claude.")


def llamar_openai(cfg, prompt, imagenes):
    modelo = cfg["modelo"] or MODELOS_POR_DEFECTO["openai"]
    content = [{"type": "text", "text": prompt}]
    for b64, mime in imagenes:
        content.append({"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}})
    r = requests.post("https://api.openai.com/v1/chat/completions",
                      headers={"Authorization": f"Bearer {cfg['api_key']}", "Content-Type": "application/json"},
                      json={"model": modelo, "messages": [{"role": "user", "content": content}],
                            "response_format": {"type": "json_object"}, "temperature": 0.1}, timeout=180)
    r.raise_for_status()
    return r.json()["choices"][0]["message"]["content"]


def llamar_anthropic(cfg, prompt, imagenes):
    modelo = cfg["modelo"] or MODELOS_POR_DEFECTO["anthropic"]
    content = []
    for b64, mime in imagenes:
        content.append({"type": "image", "source": {"type": "base64", "media_type": mime, "data": b64}})
    content.append({"type": "text", "text": prompt})
    r = requests.post("https://api.anthropic.com/v1/messages",
                      headers={"x-api-key": cfg["api_key"], "anthropic-version": "2023-06-01",
                               "content-type": "application/json"},
                      json={"model": modelo, "max_tokens": 2000, "temperature": 0.1,
                            "messages": [{"role": "user", "content": content}]}, timeout=180)
    r.raise_for_status()
    return r.json()["content"][0]["text"]


def llamar_ia(cfg, prompt, imagenes=None):
    imagenes = imagenes or []
    p = cfg["proveedor"]
    if p == "gemini": return llamar_gemini(cfg, prompt, imagenes)
    if p == "openai": return llamar_openai(cfg, prompt, imagenes)
    if p == "anthropic": return llamar_anthropic(cfg, prompt, imagenes)
    raise ValueError(f"Proveedor desconocido: {p}")


def parsear_json(texto):
    texto = re.sub(r"^```(json)?", "", texto.strip()).strip()
    texto = re.sub(r"```$", "", texto).strip()
    try:
        return json.loads(texto)
    except json.JSONDecodeError:
        m = re.search(r"\{.*\}", texto, re.DOTALL)
        if m:
            return json.loads(m.group(0))
        raise


# ---------------------------------------------------------------------------
# Enriquecimiento gratis (Google Books / OpenLibrary)
# ---------------------------------------------------------------------------

IDIOMAS = {"es": "Español", "en": "Inglés", "it": "Italiano", "ru": "Ruso", "fr": "Francés",
           "de": "Alemán", "pt": "Portugués", "la": "Latín", "el": "Griego"}


def google_books(q):
    try:
        r = requests.get("https://www.googleapis.com/books/v1/volumes",
                         params={"q": q, "maxResults": 1}, timeout=30)
        items = r.json().get("items")
        if not items:
            return {}
        v = items[0]["volumeInfo"]
        isbn = ""
        for x in v.get("industryIdentifiers", []):
            if "ISBN" in x.get("type", ""):
                isbn = x.get("identifier", ""); break
        return {"titulo_libro": v.get("title", ""), "subtitulo": v.get("subtitle", ""),
                "autor": ", ".join(v.get("authors", [])), "editorial": v.get("publisher", ""),
                "anio": (v.get("publishedDate", "") or "")[:4],
                "idioma": IDIOMAS.get(v.get("language", ""), v.get("language", "")),
                "paginas": str(v.get("pageCount", "") or ""), "tema_genero": ", ".join(v.get("categories", [])),
                "descripcion": v.get("description", ""), "isbn": isbn}
    except Exception:
        return {}


def openlibrary_isbn(isbn):
    try:
        r = requests.get("https://openlibrary.org/api/books",
                         params={"format": "json", "jscmd": "data", "bibkeys": f"ISBN:{isbn}"}, timeout=30)
        v = r.json().get(f"ISBN:{isbn}")
        if not v:
            return {}
        anio = ""
        m = re.search(r"\d{4}", v.get("publish_date", "") or "")
        if m: anio = m.group(0)
        return {"titulo_libro": v.get("title", ""), "subtitulo": v.get("subtitle", ""),
                "autor": ", ".join(a.get("name", "") for a in v.get("authors", [])),
                "editorial": ", ".join(p.get("name", "") for p in v.get("publishers", [])),
                "anio": anio, "paginas": str(v.get("number_of_pages", "") or ""),
                "ciudad": ", ".join(p.get("name", "") for p in v.get("publish_places", []))}
    except Exception:
        return {}


def openlibrary_buscar(titulo, autor):
    try:
        params = {"limit": 1, "title": titulo,
                  "fields": "title,author_name,first_publish_year,publisher,number_of_pages_median"}
        if autor:
            params["author"] = autor
        r = requests.get("https://openlibrary.org/search.json", params=params, timeout=30)
        doc = (r.json().get("docs") or [None])[0]
        if not doc:
            return {}
        return {"paginas": str(doc.get("number_of_pages_median", "") or ""),
                "editorial": (doc.get("publisher") or [""])[0],
                "anio": str(doc.get("first_publish_year", "") or "")}
    except Exception:
        return {}


def rellenar(datos, extra):
    for k, v in extra.items():
        if v and not limpiar(datos.get(k)):
            datos[k] = v


def enriquecer(datos):
    isbn = re.sub(r"[^0-9Xx]", "", datos.get("isbn", "") or "")
    if len(isbn) in (10, 13):
        rellenar(datos, google_books(f"isbn:{isbn}"))
        rellenar(datos, openlibrary_isbn(isbn))
    # Búsqueda por título (completa páginas/editorial/año en libros sin ISBN).
    if limpiar(datos.get("titulo_libro")):
        if not all(limpiar(datos.get(k)) for k in ("paginas", "editorial", "anio")):
            q = "intitle:" + datos["titulo_libro"]
            if limpiar(datos.get("autor")):
                q += " inauthor:" + datos["autor"]
            rellenar(datos, google_books(q))
        if not limpiar(datos.get("paginas")):
            rellenar(datos, openlibrary_buscar(datos["titulo_libro"], datos.get("autor", "")))
    return datos


# ---------------------------------------------------------------------------
# Armado del registro (61 columnas)
# ---------------------------------------------------------------------------

def limpiar(x):
    return "" if x is None else str(x).strip()


def armar_titulo(datos, cfg):
    partes = [limpiar(datos.get("titulo_libro")), limpiar(datos.get("autor")), limpiar(cfg.get("etiqueta_titulo"))]
    t = re.sub(r"\s+", " ", " ".join(p for p in partes if p)).strip()
    maxl = int(cfg.get("max_largo_titulo", 60))
    return t[:maxl].rstrip() if len(t) > maxl else t


def resolver_precio(datos, cfg):
    modo = cfg.get("precio", "manual")
    if isinstance(modo, (int, float)):
        return int(modo)
    modo = str(modo).strip().lower()
    if modo == "sugerir":
        a, b = datos.get("precio_sugerido_min"), datos.get("precio_sugerido_max")
        try:
            a = int(a) if a else None
        except (TypeError, ValueError):
            a = None
        try:
            b = int(b) if b else None
        except (TypeError, ValueError):
            b = None
        if a and b: return round((a + b) / 2)
        return a or b or ""
    if modo.isdigit():
        return int(modo)
    return ""


def _num(x):
    try:
        return float(str(x).replace(",", "."))
    except (TypeError, ValueError):
        return None


def _italica(s):
    out = []
    for ch in str(s):
        o = ord(ch)
        if ch == "h": out.append("ℎ")
        elif 65 <= o <= 90: out.append(chr(0x1D434 + (o - 65)))
        elif 97 <= o <= 122: out.append(chr(0x1D44E + (o - 97)))
        elif 48 <= o <= 57: out.append(chr(0x1D7F6 + (o - 48)))
        else: out.append(ch)
    return "".join(out)


def componer_descripcion(datos, cfg):
    """1) ficha bibliográfica, 2) texto de la IA, 3) datos de la librería."""
    titulo = limpiar(datos.get("titulo_libro"))
    titulo_fmt = _italica(titulo) if cfg.get("titulo_italica") and titulo else titulo
    b1 = ", ".join(x for x in [limpiar(datos.get("autor")), titulo_fmt] if x)
    b2 = ", ".join(x for x in [limpiar(datos.get("ciudad")), limpiar(datos.get("editorial")),
                               limpiar(datos.get("anio"))] if x)
    biblio = ", ".join(x for x in [b1, b2] if x)
    if biblio:
        biblio += "."
    extras = []
    if limpiar(datos.get("paginas")):
        extras.append(limpiar(datos.get("paginas")) + " páginas.")
    est = limpiar(datos.get("estado"))
    if est:
        extras.append(est if est[-1] in ".!?" else est + ".")
    bloque1 = " ".join(x for x in [biblio, " ".join(extras)] if x).strip()
    prosa = limpiar(datos.get("descripcion"))
    cierre = "\n\n".join(x for x in [limpiar(cfg.get("retiro_texto")), limpiar(cfg.get("libreria"))] if x)
    return "\n\n".join(x for x in [bloque1, prosa, cierre] if x)


def construir_registro(datos, cfg):
    import math
    alt, anc, gro, pes = (_num(datos.get(k)) for k in ("altura_cm", "ancho_cm", "grosor_cm", "peso_g"))
    derivados = {
        "__titulo": armar_titulo(datos, cfg),
        "__condicion": limpiar(datos.get("condicion")) or cfg.get("condicion", "Usado"),
        "__sku": cfg.get("sku_por_defecto", ""),
        "__stock": cfg.get("stock", 1),
        "__precio": resolver_precio(datos, cfg),
        "__descripcion": componer_descripcion(datos, cfg),
        # Medidas del paquete: medida del libro + margen si está; si no, el default de config.
        "__paq_ancho": str(math.ceil(anc) + 2) if anc else cfg.get("paq_ancho", ""),
        "__paq_alto": str(math.ceil(alt) + 2) if alt else cfg.get("paq_alto", ""),
        "__paq_prof": str(math.ceil(gro) + 1) if gro else cfg.get("paq_prof", ""),
        "__paq_peso": str(round(pes / 1000 + 0.1, 3)) if pes else cfg.get("paq_peso", ""),
        "__cuotas": cfg.get("cuotas", ""), "__costo_cuotas": cfg.get("costo_cuotas", ""),
        "__forma_envio": cfg.get("forma_envio", ""), "__costo_envio": cfg.get("costo_envio", ""),
        "__retiro": cfg.get("retiro", ""),
        "__u_altura": "cm" if limpiar(datos.get("altura_cm")) else "",
        "__u_ancho": "cm" if limpiar(datos.get("ancho_cm")) else "",
        "__u_peso": "g" if limpiar(datos.get("peso_g")) else "",
    }
    reg = OrderedDict()
    for header, clave in COLUMNAS:
        if clave is None:
            reg[header] = ""
        elif clave.startswith("__"):
            reg[header] = derivados.get(clave, "")
        else:
            reg[header] = limpiar(datos.get(clave))
    return reg


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def escribir_excel(registros, ruta_salida):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Publicaciones"
    headers = [h for h, _ in COLUMNAS]
    fill = PatternFill("solid", fgColor="FFE699")
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True); c.fill = fill; c.alignment = Alignment(horizontal="center")
    for reg in registros:
        ws.append([reg.get(h, "") for h in headers])
    for i, h in enumerate(headers, start=1):
        ancho = 40 if h in ("Título", "Descripción", "Título del libro") else 16
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
    ws.freeze_panes = "A2"
    wb.save(ruta_salida)


def escribir_en_plantilla(registros, plantilla, salida):
    """Rellena la planilla oficial de ML ('Publicar varios productos') conservando su estructura."""
    wb = openpyxl.load_workbook(plantilla)
    if "Libros Físicos" not in wb.sheetnames:
        raise RuntimeError("La plantilla no tiene la hoja 'Libros Físicos'. "
                           "¿Es la planilla de 'Publicar varios productos' > Libros Físicos?")
    ws = wb["Libros Físicos"]
    PRIMERA = 9  # primera fila de datos
    # Última fila de ejemplo (col D = Condición con valor).
    ultima = PRIMERA - 1
    r = PRIMERA
    while ws.cell(r, 4).value not in (None, ""):
        ultima = r
        r += 1
    headers = [h for h, _ in COLUMNAS]
    # Escribir cada libro (solo valores no vacíos: así se respetan los defaults grises de la planilla).
    for i, reg in enumerate(registros):
        fila = PRIMERA + i
        for col, h in enumerate(headers, start=1):
            val = reg.get(h, "")
            if val not in ("", None):
                ws.cell(fila, col).value = val
    # Limpiar filas de ejemplo sobrantes para que ML no intente publicarlas.
    for fila in range(PRIMERA + len(registros), ultima + 1):
        for col in range(1, len(headers) + 1):
            ws.cell(fila, col).value = None
    wb.save(salida)


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

def cargar_config(ruta):
    cfg = dict(CONFIG_POR_DEFECTO)
    if ruta.exists():
        try:
            with open(ruta, "r", encoding="utf-8") as f:
                data = json.load(f)
            for k, v in data.items():
                if v != "" or k in ("etiqueta_titulo", "sku_por_defecto"):
                    cfg[k] = v
        except json.JSONDecodeError as e:
            print(f"⚠️  config.json inválido ({e}). Uso valores por defecto.")
    if not cfg.get("api_key"):
        cfg["api_key"] = os.environ.get(VAR_ENTORNO_KEY.get(cfg["proveedor"], ""), "")
    return cfg


# ---------------------------------------------------------------------------
# Agrupar fotos por libro (prefijo + doble guion bajo)
# ---------------------------------------------------------------------------

def agrupar_fotos(fotos):
    """Devuelve lista de (nombre_grupo, [rutas]). 'libro__a.jpg' y 'libro__b.jpg' -> mismo grupo."""
    grupos = OrderedDict()
    for f in fotos:
        stem = f.stem
        clave = stem.split("__")[0] if "__" in stem else stem
        grupos.setdefault(clave, []).append(f)
    return list(grupos.items())


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Extrae datos de libros para la planilla 'Publicar' de Mercado Libre.")
    ap.add_argument("--config", default=str(RAIZ / "config.json"))
    ap.add_argument("--carpeta")
    ap.add_argument("--salida")
    ap.add_argument("--texto", help="Identificar un libro por texto (título/autor)")
    ap.add_argument("--isbn", help="Identificar un libro por ISBN")
    ap.add_argument("--plantilla", help="Ruta a la planilla oficial de ML para rellenarla (en vez de crear una nueva)")
    args = ap.parse_args()

    cfg = cargar_config(Path(args.config))
    if args.carpeta: cfg["carpeta_fotos"] = args.carpeta
    if args.salida: cfg["archivo_salida"] = args.salida
    if args.plantilla: cfg["plantilla"] = args.plantilla

    if cfg["proveedor"] not in MODELOS_POR_DEFECTO:
        print(f"❌ Proveedor inválido: '{cfg['proveedor']}'. Usá gemini, openai o anthropic."); sys.exit(1)

    necesita_key = not (args.isbn and not args.texto)
    if not cfg.get("api_key") and necesita_key:
        print(f"❌ Falta la API key. Poné 'api_key' en config.json o la variable "
              f"{VAR_ENTORNO_KEY[cfg['proveedor']]}. Ver README.md."); sys.exit(1)

    registros = []

    if args.texto or args.isbn:
        try:
            if args.isbn:
                print(f"🔎 ISBN {args.isbn} ...")
                datos = {"isbn": re.sub(r'[^0-9Xx]', '', args.isbn), "confianza": "alta"}
                enriquecer(datos)
                if not limpiar(datos.get("titulo_libro")):
                    datos = parsear_json(llamar_ia(cfg, prompt_texto("ISBN " + args.isbn))); enriquecer(datos)
            else:
                print(f"🔎 {args.texto} ...")
                datos = parsear_json(llamar_ia(cfg, prompt_texto(args.texto))); enriquecer(datos)
            registros.append(construir_registro(datos, cfg))
            print(f"   ✔ {registros[-1]['Título']}")
        except Exception as e:
            print(f"   ✖ Error: {e}")
    else:
        carpeta = Path(cfg["carpeta_fotos"])
        if not carpeta.is_absolute():
            carpeta = RAIZ / carpeta
        if not carpeta.exists():
            print(f"❌ No existe la carpeta de fotos: {carpeta}"); sys.exit(1)
        fotos = sorted(p for p in carpeta.iterdir()
                       if p.is_file() and p.suffix.lower() in EXTENSIONES_IMAGEN)
        if not fotos:
            print(f"❌ No encontré fotos en {carpeta}"); sys.exit(1)

        grupos = agrupar_fotos(fotos)
        print(f"📚 {len(fotos)} foto(s) → {len(grupos)} libro(s) con "
              f"{cfg['proveedor']}\n")
        for i, (nombre, rutas) in enumerate(grupos, start=1):
            etiqueta = nombre + (f" ({len(rutas)} fotos)" if len(rutas) > 1 else "")
            print(f"[{i}/{len(grupos)}] {etiqueta} ...", end=" ", flush=True)
            intentos = 0
            while True:
                try:
                    imagenes = [imagen_a_b64(r) for r in rutas]
                    datos = parsear_json(llamar_ia(cfg, prompt_texto(""), imagenes))
                    enriquecer(datos)
                    reg = construir_registro(datos, cfg)
                    registros.append(reg)
                    conf = limpiar(datos.get("confianza")) or "s/d"
                    print(f"✔ {reg['Título']}  (confianza: {conf})")
                    break
                except requests.HTTPError as e:
                    intentos += 1
                    code = e.response.status_code if e.response is not None else "?"
                    if code == 429 and intentos <= 4:
                        espera = 2 ** intentos
                        print(f"⏳ 429, reintento en {espera}s ...", end=" ", flush=True)
                        time.sleep(espera); continue
                    print(f"✖ Error HTTP {code}")
                    registros.append(construir_registro({"observaciones": f"Error HTTP {code}"}, cfg))
                    break
                except Exception as e:
                    print(f"✖ Error: {e}")
                    registros.append(construir_registro({"observaciones": f"Error: {e}"}, cfg))
                    break

    if not registros:
        print("\nNo se generó ninguna fila."); sys.exit(1)

    salida = Path(cfg["archivo_salida"])
    if not salida.is_absolute():
        salida = RAIZ / salida

    plantilla = limpiar(cfg.get("plantilla"))
    if plantilla:
        pl = Path(plantilla)
        if not pl.is_absolute():
            pl = RAIZ / pl
        if not pl.exists():
            print(f"❌ No encuentro la plantilla: {pl}"); sys.exit(1)
        try:
            escribir_en_plantilla(registros, pl, salida)
            print(f"\n✅ Listo. Planilla de ML rellenada con {len(registros)} libro(s):\n   {salida}")
            print("   Es la misma planilla, lista para subir a Mercado Libre. Completá Fotos y, si falta, el Precio.")
        except Exception as e:
            print(f"❌ No pude rellenar la plantilla ({e}). Genero la salida normal.")
            escribir_excel(registros, salida)
    else:
        escribir_excel(registros, salida)
        print(f"\n✅ Listo. {len(registros)} fila(s) con las 61 columnas de ML en:\n   {salida}")
        print("   Abrí el archivo, copiá las filas y pegalas en la planilla que bajás de Mercado Libre.")


if __name__ == "__main__":
    main()
