import base64
import io
import json
import os
import re
import traceback
from pathlib import Path

import anthropic
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

app = FastAPI(title="Completador de Planilla ML - Libros Físicos")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

FIXED_VALUES = {
    "condición": "Usado",
    "stock": 1,
    "descripción": "Librería Los Siete Pilares / Buen estado. Se retira por zona de Paraguay y Reconquista",
    "forma de envío": "Mercado Envíos | Mercado Envío Flex",
    "costo de envío": "A cargo del comprador",
    "retiro en persona": "Acepto",
    "tipo de garantía": "Garantía del vendedor",
    "tiempo de garantía": 1,
    "unidad de tiempo de garantía": "meses",
}

SKIP_COLUMNS_KEYWORDS = [
    "cargo por vender", "cuotas", "costo por ofrecer cuotas",
    "factura a", "índice", "fotos", "sku", "precio",
]


def normalize(text: str) -> str:
    return text.lower().strip()


def truncate_title(title: str, max_len: int = 60) -> str:
    if len(title) <= max_len:
        return title
    truncated = title[:max_len]
    last_space = truncated.rfind(" ")
    if last_space > 0:
        truncated = truncated[:last_space]
    return truncated.strip()


def build_ml_title(book_title: str, author: str) -> str:
    suffix = " Microcentro"
    raw = f"{book_title} {author}{suffix}"
    return truncate_title(raw, 60)


def find_header_row(ws) -> tuple[int | None, dict[str, int]]:
    """Search rows 1-10 for the header row; return (row_number, {col_name: col_index})."""
    for row_idx in range(1, 11):
        row = ws[row_idx]
        headers = {}
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                headers[normalize(cell.value)] = cell.column
        if len(headers) >= 5:
            return row_idx, headers
    return None, {}


def match_column(headers: dict[str, int], keywords: list[str]) -> int | None:
    for key in headers:
        for kw in keywords:
            if kw in key:
                return headers[key]
    return None


def extract_book_data_from_image(client: anthropic.Anthropic, image_b64: str, media_type: str) -> dict:
    prompt = """Analizá esta imagen de un libro y extraé los siguientes datos en formato JSON.
Devolvé SOLO el JSON, sin texto adicional, con exactamente estas claves:

{
  "titulo_libro": "título exacto del libro",
  "autor": "nombre completo del autor (Apellido Nombre si está visible)",
  "editorial": "editorial si es visible, sino null",
  "subtitulo": "subtítulo si es visible, sino null",
  "serie": "serie o colección si es visible, sino null",
  "edicion": "número de edición si es visible, sino null",
  "volumen": "número de volumen si es visible, sino null",
  "anio_publicacion": "año de publicación si es visible, sino null",
  "isbn": "ISBN completo si es visible (solo dígitos y guiones), sino null",
  "idioma": "idioma del libro (Español, Inglés, Francés, etc.)",
  "tapa": "Dura o Blanda según el tipo de encuadernación visible",
  "ancho_cm": número entero estimado del ancho en cm,
  "alto_cm": número entero estimado del alto en cm,
  "profundidad_cm": número entero estimado de la profundidad (grosor) en cm,
  "peso_kg": número entero estimado del peso en kg (1 para la mayoría de libros estándar)
}

REGLAS:
- Para el autor: usar el formato que aparece en el libro.
- ISBN: solo si es CLARAMENTE visible en la imagen. NO inventar.
- Medidas: estimar razonablemente. Un libro de bolsillo típico: 19x13x2cm. Un libro estándar: 23x16x3cm. Un libro grande: 28x22x4cm.
- Peso: normalmente 1 kg para libros estándar, hasta 2 kg para libros grandes/gruesos.
- Si algún dato no es visible ni inferible, usar null.
- Responder SOLO con el JSON, sin markdown, sin explicaciones."""

    response = client.messages.create(
        model="claude-opus-4-8",
        max_tokens=1024,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": image_b64,
                        },
                    },
                    {"type": "text", "text": prompt},
                ],
            }
        ],
    )

    raw = response.content[0].text.strip()
    raw = re.sub(r"^```json\s*", "", raw)
    raw = re.sub(r"^```\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    return json.loads(raw)


def write_book_to_row(ws, row_num: int, headers: dict[str, int], book: dict):
    def set_cell(col_idx, value):
        if col_idx is None or value is None:
            return
        cell = ws.cell(row=row_num, column=col_idx)
        # Preserve any existing data validation by not touching cells we shouldn't
        cell.value = value

    # Build ML title
    titulo = book.get("titulo_libro") or ""
    autor = book.get("autor") or ""
    ml_title = build_ml_title(titulo, autor) if titulo else None

    col_map = {
        # Title column (long name with "incluí producto")
        "título": match_column(headers, ["título:", "incluí producto"]),
        "condición": match_column(headers, ["condición", "condicion"]),
        "isbn": match_column(headers, ["isbn"]),
        "stock": match_column(headers, ["stock"]),
        "descripción": match_column(headers, ["descripción", "descripcion"]),
        "ancho": match_column(headers, ["ancho"]),
        "alto": match_column(headers, ["alto"]),
        "profundidad": match_column(headers, ["profundidad"]),
        "peso": match_column(headers, ["peso"]),
        "forma de envío": match_column(headers, ["forma de envío", "forma de envio"]),
        "costo de envío": match_column(headers, ["costo de envío", "costo de envio"]),
        "retiro en persona": match_column(headers, ["retiro en persona"]),
        "tipo de garantía": match_column(headers, ["tipo de garantía", "tipo de garantia"]),
        "tiempo de garantía": match_column(headers, ["tiempo de garantía", "tiempo de garantia"]),
        "unidad de tiempo": match_column(headers, ["unidad de tiempo"]),
        "titulo_libro": match_column(headers, ["título del libro", "titulo del libro"]),
        "autor": match_column(headers, ["autor"]),
        "editorial": match_column(headers, ["editorial"]),
        "subtitulo": match_column(headers, ["subtítulo", "subtitulo"]),
        "serie": match_column(headers, ["serie"]),
        "edicion": match_column(headers, ["edición", "edicion"]),
        "volumen": match_column(headers, ["volumen"]),
        "anio": match_column(headers, ["año de publicación", "ano de publicacion"]),
        "idioma": match_column(headers, ["idioma"]),
        "tapa": match_column(headers, ["tapa del libro", "tapa"]),
    }

    set_cell(col_map["título"], ml_title)
    set_cell(col_map["condición"], "Usado")
    set_cell(col_map["isbn"], book.get("isbn"))
    set_cell(col_map["stock"], 1)
    set_cell(col_map["descripción"], FIXED_VALUES["descripción"])
    set_cell(col_map["ancho"], book.get("ancho_cm"))
    set_cell(col_map["alto"], book.get("alto_cm"))
    set_cell(col_map["profundidad"], book.get("profundidad_cm"))
    set_cell(col_map["peso"], book.get("peso_kg"))
    set_cell(col_map["forma de envío"], FIXED_VALUES["forma de envío"])
    set_cell(col_map["costo de envío"], FIXED_VALUES["costo de envío"])
    set_cell(col_map["retiro en persona"], "Acepto")
    set_cell(col_map["tipo de garantía"], FIXED_VALUES["tipo de garantía"])
    set_cell(col_map["tiempo de garantía"], 1)
    set_cell(col_map["unidad de tiempo"], "meses")
    set_cell(col_map["titulo_libro"], titulo or None)
    set_cell(col_map["autor"], autor or None)
    set_cell(col_map["editorial"], book.get("editorial"))
    set_cell(col_map["subtitulo"], book.get("subtitulo"))
    set_cell(col_map["serie"], book.get("serie"))
    set_cell(col_map["edicion"], book.get("edicion"))
    set_cell(col_map["volumen"], book.get("volumen"))
    set_cell(col_map["anio"], book.get("anio_publicacion"))
    set_cell(col_map["idioma"], book.get("idioma") or "Español")
    set_cell(col_map["tapa"], book.get("tapa"))


@app.get("/", response_class=HTMLResponse)
async def root():
    html_path = Path(__file__).parent / "index.html"
    return html_path.read_text(encoding="utf-8")


@app.post("/process")
async def process(
    excel_file: UploadFile = File(...),
    images: list[UploadFile] = File(...),
    api_key: str = Form(...),
):
    if not api_key.strip():
        raise HTTPException(status_code=400, detail="API key de Anthropic requerida.")

    client = anthropic.Anthropic(api_key=api_key.strip())

    # Load workbook preserving formulas and structure
    excel_bytes = await excel_file.read()
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active

    header_row, headers = find_header_row(ws)
    if header_row is None:
        raise HTTPException(status_code=422, detail="No se encontró fila de encabezados en el Excel.")

    data_start_row = max(8, header_row + 1)

    books = []
    errors = []

    for idx, img_file in enumerate(images):
        try:
            img_bytes = await img_file.read()
            img_b64 = base64.standard_b64encode(img_bytes).decode()
            content_type = img_file.content_type or "image/jpeg"
            # Normalize media type
            if "png" in content_type:
                media_type = "image/png"
            elif "gif" in content_type:
                media_type = "image/gif"
            elif "webp" in content_type:
                media_type = "image/webp"
            else:
                media_type = "image/jpeg"

            book_data = extract_book_data_from_image(client, img_b64, media_type)
            books.append({"file": img_file.filename, "data": book_data, "row": data_start_row + idx})
        except Exception as e:
            errors.append({"file": img_file.filename, "error": str(e)})

    for book_info in books:
        write_book_to_row(ws, book_info["row"], headers, book_info["data"])

    # Save workbook to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    original_name = excel_file.filename or "planilla.xlsx"
    stem = original_name.rsplit(".", 1)[0]
    output_name = f"{stem}_completada.xlsx"

    output_path = Path("/tmp") / output_name
    output_path.write_bytes(output.read())

    return {
        "status": "ok",
        "books_processed": len(books),
        "errors": errors,
        "books": [{"file": b["file"], "row": b["row"], "titulo": b["data"].get("titulo_libro"), "autor": b["data"].get("autor")} for b in books],
        "download_path": f"/download/{output_name}",
    }


@app.get("/download/{filename}")
async def download(filename: str):
    # Sanitize filename - only allow alphanumeric, dash, underscore, dot
    safe = re.sub(r"[^a-zA-Z0-9_\-\. ]", "", filename)
    path = Path("/tmp") / safe
    if not path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado.")
    return FileResponse(
        path=str(path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=safe,
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
